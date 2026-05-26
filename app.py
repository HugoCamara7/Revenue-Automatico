from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl
import streamlit as st

from generar_matrixify_descuentos import analyze_discount_preview, build_discount_workbook, extract_input_lookup_keys


SITES = {
    "Rockford.pe": {
        "brand": "COLUMBIA, ROCKFORD, PATAGONIA, SOREL, MOUNTAIN HARDWEAR",
        "brands": ["COLUMBIA", "ROCKFORD", "PATAGONIA", "SOREL", "MOUNTAIN HARDWEAR"],
        "vendor": "rockfordpe",
        "output": "matrixify_revenue_rockford.xlsx",
    },
    "Columbia.pe": {
        "brand": "COLUMBIA",
        "brands": ["COLUMBIA"],
        "vendor": "columbiape",
        "output": "matrixify_revenue_columbia.xlsx",
    },
    "Hushpuppies.pe": {
        "brand": "HUSH PUPPIES",
        "brands": ["HUSH PUPPIES"],
        "vendor": "hushpuppiespe",
        "output": "matrixify_revenue_hushpuppies.xlsx",
    },
    "Vans.pe": {
        "brand": "VANS",
        "brands": ["VANS"],
        "vendor": "vanspe",
        "output": "matrixify_revenue_vans.xlsx",
    },
    "Supermall.pe": {
        "brand": "MULTIMARCA",
        "brands": ["COLUMBIA", "HUSH PUPPIES", "ROCKFORD", "PATAGONIA", "SOREL", "MOUNTAIN HARDWEAR", "VANS"],
        "vendor": "supermallpe",
        "output": "matrixify_revenue_supermall.xlsx",
    },
}


st.set_page_config(page_title="Matrixify Descuentos", layout="wide")

st.markdown(
    """
    <style>
    .stApp {
        background: #ffffff;
        color: #001f4f;
    }
    [data-testid="stSidebar"] {
        background: #eef2f7;
    }
    .main-block {
        max-width: 960px;
        margin: 0 auto;
        padding-top: 28px;
    }
    .brand-logo {
        font-size: 58px;
        line-height: 1;
        font-weight: 800;
        color: #15329b;
        letter-spacing: 1px;
    }
    .brand-subtitle {
        color: #15329b;
        font-size: 15px;
        font-weight: 700;
        letter-spacing: 7px;
        margin-bottom: 64px;
    }
    .shopify-badge {
        height: 84px;
        width: 84px;
        border-radius: 12px;
        background: #95bf47;
        color: #ffffff;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 48px;
        font-weight: 800;
        transform: rotate(4deg);
        margin-left: auto;
    }
    .info-box {
        background: #eef6ff;
        border: 1px solid #c3ddff;
        border-radius: 8px;
        padding: 20px 24px;
        margin: 34px 0 40px 0;
    }
    .section-card {
        border: 1px solid #d7e1ef;
        border-radius: 8px;
        padding: 28px 24px;
        margin-bottom: 18px;
    }
    .upload-card {
        border: 1px dashed #8ab4ff;
        border-radius: 8px;
        padding: 20px 18px;
        margin-bottom: 16px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def save_upload(uploaded_file, folder: Path) -> Path:
    path = folder / uploaded_file.name
    path.write_bytes(uploaded_file.getbuffer())
    return path


def read_summary(output_path: Path) -> tuple[list[str], list[list[object]], int]:
    workbook = openpyxl.load_workbook(output_path, data_only=True, read_only=True)
    summary = workbook["Resumen"]
    headers = [summary.cell(row=1, column=col).value for col in range(1, summary.max_column + 1)]
    rows = [
        [summary.cell(row=row, column=col).value for col in range(1, summary.max_column + 1)]
        for row in range(2, summary.max_row + 1)
    ]
    missing_count = workbook["No encontrados"].max_row - 1 if "No encontrados" in workbook.sheetnames else 0
    workbook.close()
    return headers, rows, missing_count


def validate_matrixify_site(matrixify_path: Path, expected_vendor: str) -> tuple[bool, str]:
    workbook = openpyxl.load_workbook(matrixify_path, data_only=True, read_only=True)
    try:
        if "Export Summary" in workbook.sheetnames:
            summary = workbook["Export Summary"]
            headers = [summary.cell(row=1, column=col).value for col in range(1, summary.max_column + 1)]
            header_map = {str(value).strip(): index + 1 for index, value in enumerate(headers) if value}
            domain_col = header_map.get("Shopify Domain")
            if domain_col:
                domain = str(summary.cell(row=2, column=domain_col).value or "").strip().lower()
                expected = expected_vendor.lower()
                if domain and expected not in domain:
                    return (
                        False,
                        f"El Matrixify cargado parece ser de `{domain}`, pero elegiste `{expected_vendor}`. "
                        "Sube el ultimo catalogo Matrixify del mismo sitio destino.",
                    )
                return True, ""

        products = workbook["Products"] if "Products" in workbook.sheetnames else workbook.active
        headers = [products.cell(row=1, column=col).value for col in range(1, products.max_column + 1)]
        header_map = {str(value).strip(): index + 1 for index, value in enumerate(headers) if value}
        vendor_col = header_map.get("Vendor")
        if vendor_col:
            vendors = set()
            for row in range(2, min(products.max_row, 500) + 1):
                vendor = str(products.cell(row=row, column=vendor_col).value or "").strip().lower()
                if vendor:
                    vendors.add(vendor)
            expected = expected_vendor.lower()
            wrong_vendors = sorted(vendor for vendor in vendors if vendor != expected)
            if wrong_vendors and expected not in vendors:
                return (
                    False,
                    f"El Matrixify cargado tiene Vendor `{', '.join(wrong_vendors[:5])}`, "
                    f"pero elegiste `{expected_vendor}`. Sube el catalogo Matrixify correcto del sitio destino.",
                )
            if vendors and expected not in vendors:
                return (
                    False,
                    f"No encontre Vendor `{expected_vendor}` en el Matrixify cargado. "
                    f"Vendors encontrados: `{', '.join(sorted(vendors)[:5])}`.",
                )
    finally:
        workbook.close()

    return True, "No pude validar el sitio porque el Matrixify no trae `Export Summary`."


def normalize_column_name(value: str) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def find_bigquery_column(available_columns: list[str], candidates: list[str]) -> str:
    normalized = {normalize_column_name(column): column for column in available_columns}
    for candidate in candidates:
        found = normalized.get(normalize_column_name(candidate))
        if found:
            return found
    return ""


def bigquery_field_expression(column_name: str) -> str:
    return f"CAST(`{column_name}` AS STRING)"


def resolve_bigquery_columns(client, table_id: str, config: dict) -> tuple[str, str, str, str]:
    table_schema = client.get_table(table_id).schema
    available_columns = [field.name for field in table_schema]

    id_column = str(config.get("id_column", "")).strip() or find_bigquery_column(
        available_columns,
        ["CODINT_MA", "ID_PRODUCTO", "ID PRODUCTO", "SKU", "Variant SKU", "sku_producto"],
    )
    brand_column = str(config.get("brand_column", "")).strip() or find_bigquery_column(
        available_columns,
        ["MARCA_MA", "MARCA", "BRAND", "VENDOR"],
    )
    modcol_column = str(config.get("modcol_column", "")).strip() or find_bigquery_column(
        available_columns,
        ["COD MOD COL", "COD_MOD_COL", "MODCOL", "MOD_COL", "Mod-Col", "codmod_codcol_ma"],
    )

    if modcol_column:
        modcol_expression = bigquery_field_expression(modcol_column)
    else:
        model_column = str(config.get("model_column", "")).strip() or find_bigquery_column(
            available_columns,
            ["CODMOD_MA", "CODMOD", "COD_MODELO", "CODMODELO", "MODELO"],
        )
        color_column = str(config.get("color_column", "")).strip() or find_bigquery_column(
            available_columns,
            ["CODCOL_MA", "CODCOL", "COD_COLOR", "CODCOLOR", "COLOR"],
        )
        if model_column and color_column:
            modcol_expression = f"CONCAT({bigquery_field_expression(model_column)}, '-', {bigquery_field_expression(color_column)})"
            modcol_column = "__MODEL_COLOR__"
        else:
            modcol_expression = "CAST(NULL AS STRING)"

    missing = []
    if not id_column:
        missing.append("ID producto")
    if not brand_column:
        missing.append("Marca")
    if not modcol_column:
        missing.append("MODCOL o CODMOD_MA + CODCOL_MA")
    if missing:
        raise RuntimeError(
            "No pude detectar columnas necesarias en BigQuery: "
            f"{', '.join(missing)}. Columnas disponibles: {', '.join(available_columns[:80])}"
        )

    return id_column, modcol_expression, brand_column, ", ".join([id_column, modcol_column, brand_column])


@st.cache_data(ttl=3600, show_spinner=False)
def load_brand_lookup_from_bigquery(
    ids: tuple[str, ...],
    modcols: tuple[str, ...],
    selected_brands: tuple[str, ...],
) -> dict[str, str]:
    from google.cloud import bigquery
    from google.oauth2 import service_account

    config = {}
    try:
        if "bigquery" in st.secrets:
            config.update(dict(st.secrets["bigquery"]))
        if "gcp_service_account" in st.secrets:
            config["service_account_info"] = dict(st.secrets["gcp_service_account"])
    except Exception:
        return {}

    service_account_json = config.pop("service_account_json", None)
    if service_account_json and "service_account_info" not in config:
        config["service_account_info"] = json.loads(service_account_json)

    enabled = str(config.get("enabled", "true")).strip().lower()
    if enabled in ("0", "false", "no", "off"):
        return {}

    credentials_info = config.get("service_account_info")
    credentials = service_account.Credentials.from_service_account_info(dict(credentials_info)) if credentials_info else None
    project_id = str(config.get("project_id") or (credentials.project_id if credentials else "")).strip()
    job_project_id = str(config.get("job_project_id") or project_id).strip() or None
    client = bigquery.Client(project=job_project_id, credentials=credentials)

    query = str(config.get("query", "")).strip().rstrip(";")
    table_id = ""
    if not query:
        table = str(config.get("table", "")).strip()
        dataset = str(config.get("dataset", "")).strip()
        if not table:
            return {}
        table_id = table if table.count(".") == 2 else f"{project_id}.{dataset}.{table}"
        id_column, modcol_expression, brand_column, _detected_columns = resolve_bigquery_columns(client, table_id, config)
        base_sql = f"`{table_id}`"
    else:
        id_column = str(config.get("id_column", "CODINT_MA")).strip()
        modcol_column = str(config.get("modcol_column", "COD MOD COL")).strip()
        brand_column = str(config.get("brand_column", "MARCA_MA")).strip()
        modcol_expression = bigquery_field_expression(modcol_column)
        base_sql = f"({query})"

    brand_filter_sql = ""
    query_parameters = [
        bigquery.ArrayQueryParameter("ids", "STRING", list(ids)),
        bigquery.ArrayQueryParameter("modcols", "STRING", list(modcols)),
    ]
    if selected_brands:
        brand_filter_sql = f"AND UPPER(CAST(`{brand_column}` AS STRING)) IN UNNEST(@selected_brands)"
        query_parameters.append(
            bigquery.ArrayQueryParameter(
                "selected_brands",
                "STRING",
                [brand.upper() for brand in selected_brands],
            )
        )

    filtered_query = f"""
    SELECT
      CAST(`{id_column}` AS STRING) AS id_value,
      {modcol_expression} AS modcol_value,
      CAST(`{brand_column}` AS STRING) AS brand_value
    FROM {base_sql}
    WHERE
      (
        CAST(`{id_column}` AS STRING) IN UNNEST(@ids)
        OR {modcol_expression} IN UNNEST(@modcols)
      )
      {brand_filter_sql}
    """

    job_config = bigquery.QueryJobConfig(
        use_legacy_sql=False,
        query_parameters=query_parameters,
    )

    query_job = client.query(
        filtered_query,
        job_config=job_config,
        location=str(config.get("location", "")).strip() or None,
    )
    rows = query_job.result(timeout=int(config.get("timeout_seconds", 90)))

    lookup: dict[str, str] = {}
    for row in rows:
        row_dict = dict(row.items())
        brand = str(row_dict.get("brand_value") or "").strip()
        if not brand:
            continue
        for field in ("id_value", "modcol_value"):
            value = row_dict.get(field)
            if value not in (None, ""):
                key = "".join(ch for ch in str(value).upper() if ch.isalnum())
                lookup[key] = brand
    return lookup


with st.sidebar:
    site_name = st.selectbox("Sitio destino", list(SITES.keys()))
    site = SITES[site_name]
    st.markdown("**Marcas permitidas**")
    st.write(site["brand"])
    selected_brands = st.multiselect(
        "Marcas a afectar",
        site["brands"],
        default=site["brands"][:1],
        help="Se filtra con el maestro de BigQuery. Si no hay maestro configurado, se usa solo el alcance del input.",
    )
    st.caption(f"Vendor: {site['vendor']} | Salida: {site['output']}")

st.sidebar.info("BigQuery se cargara al generar el archivo.")


st.markdown('<div class="main-block">', unsafe_allow_html=True)
top_left, top_right = st.columns([3, 1])
with top_left:
    st.image("forus_logo.png", width=230)
with top_right:
    st.image("shopify_logo.png", width=90)

st.header(f"Matrixify {site_name} - Shopify")
st.write(
    "Sube el input comercial y el ultimo catalogo Matrixify del sitio para conservar IDs y evitar duplicados."
)

st.markdown(
    """
    <div class="info-box">
      <strong>Flujo obligatorio:</strong><br>
      Elige el sitio destino, sube el Revenue comercial y sube el ultimo catalogo Matrixify del mismo sitio.
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="section-card"><h3>Cargar archivos obligatorios</h3></div>', unsafe_allow_html=True)

st.markdown('<div class="upload-card">', unsafe_allow_html=True)
revenue_file = st.file_uploader(
    "1. Subir input comercial / Revenue",
    type=["xlsx", "xls"],
    help="Debe incluir ID PRODUCTO, MODCOL y columnas como DCTO ANT, NUEVO DCTO o DESCUENTO.",
)
st.markdown("</div>", unsafe_allow_html=True)

st.markdown('<div class="upload-card">', unsafe_allow_html=True)
matrixify_file = st.file_uploader(
    f"2. Subir ultimo catalogo Matrixify de {site_name}",
    type=["xlsx", "xls"],
    help="Debe tener las columnas Matrixify de Products, incluyendo ID, Handle, Variant SKU, Variant Price y Compare At Price.",
)
st.markdown("</div>", unsafe_allow_html=True)

with st.expander("Regla de calculo"):
    st.write(
        "Primero se busca el SKU del Revenue en `Variant SKU`. Cuando encuentra un SKU con descuento, "
        "la app aplica ese mismo descuento a todas las variantes del mismo producto Matrixify, usando el `ID` "
        "del producto como grupo modelo-color. Cada hoja incluye todo el catalogo Matrixify, no solo los SKUs con descuento."
    )
    st.write(
        "`Variant Price` queda como precio original menos descuento. "
        "`Variant Compare At Price` solo se llena con el precio original cuando hay descuento real; "
        "si no hay descuento, queda vacio."
    )

generate = st.button(
    "Generar Matrixify de descuentos",
    type="primary",
    disabled=not matrixify_file or not revenue_file,
    use_container_width=True,
)

if generate:
    with st.status("Procesando archivos...", expanded=True) as status:
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                workdir = Path(temp_dir)
                st.write("Guardando archivos cargados...")
                revenue_path = save_upload(revenue_file, workdir)
                matrixify_path = save_upload(matrixify_file, workdir)
                output_path = workdir / site["output"]

                st.write("Validando que el Matrixify corresponda al sitio destino...")
                matrixify_ok, matrixify_message = validate_matrixify_site(matrixify_path, site["vendor"])
                if not matrixify_ok:
                    st.error(matrixify_message)
                    st.stop()
                if matrixify_message:
                    st.warning(matrixify_message)

                st.write("Leyendo codigos del input comercial...")
                ids, modcols = extract_input_lookup_keys(revenue_path)

                st.write(f"Consultando BigQuery solo por {len(ids):,} IDs y {len(modcols):,} MODCOL...")
                brand_lookup = load_brand_lookup_from_bigquery(tuple(ids), tuple(modcols), tuple(selected_brands))
                if selected_brands and not brand_lookup:
                    st.error(
                        "No se pudo obtener marca desde BigQuery para los codigos del input. "
                        "No genero el archivo para evitar tocar productos de otra marca."
                    )
                    st.stop()

                st.write("Calculando vista previa comercial...")
                preview_rows, percent_rows, preview_missing, not_affected_count = analyze_discount_preview(
                    matrixify_path, revenue_path, selected_brands, brand_lookup
                )

                st.write("Generando Excel Matrixify...")
                build_discount_workbook(matrixify_path, revenue_path, output_path, selected_brands, brand_lookup)
                headers, rows, missing_count = read_summary(output_path)
                output_bytes = output_path.read_bytes()
                status.update(label="Archivo generado correctamente.", state="complete")

            st.success("Archivo generado correctamente.")
            st.subheader("Vista previa comercial")

            if preview_rows:
                total_products = sum(row["Cod MODCOL / productos afectados"] for row in preview_rows)
                total_variants = sum(row["Variantes Matrixify afectadas"] for row in preview_rows)
                col1, col2, col3 = st.columns(3)
                col1.metric("Cargas detectadas", len(preview_rows))
                col2.metric("Cod MODCOL / productos afectados", f"{total_products:,}")
                col3.metric("Variantes afectadas", f"{total_variants:,}")

                st.write("**Resumen por carga**")
                st.dataframe(preview_rows, hide_index=True, use_container_width=True)

            if percent_rows:
                st.write("**Distribucion por porcentaje de descuento**")
                st.dataframe(percent_rows, hide_index=True, use_container_width=True)

            if preview_missing:
                st.warning(
                    f"Vista previa: {preview_missing} SKUs del input no hicieron match contra Variant SKU. "
                    "El archivo igual se genera con el catalogo completo."
                )

            if not_affected_count:
                st.warning(
                    f"{not_affected_count} codigos del input no se afectaron porque BigQuery los identifica "
                    "con otra marca o sin match. Quedaron en la hoja `No afectados por marca`."
                )

            st.subheader("Resumen")
            st.dataframe([dict(zip(headers, row)) for row in rows], hide_index=True, use_container_width=True)

            if missing_count:
                st.warning(
                    f"Hay {missing_count} SKUs del Revenue que no se encontraron en Matrixify. "
                    "Quedaron documentados en la hoja `No encontrados`."
                )

            st.download_button(
                "Descargar Matrixify generado",
                data=output_bytes,
                file_name=site["output"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as exc:
            st.error(f"No pude generar el archivo: {exc}")

st.markdown("</div>", unsafe_allow_html=True)
