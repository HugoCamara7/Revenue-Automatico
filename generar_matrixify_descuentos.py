from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


MATRIXIFY_COLUMNS = [
    "ID",
    "Handle",
    "Command",
    "Variant Inventory Item ID",
    "Variant ID",
    "Variant Command",
    "Variant SKU",
    "Variant Price",
    "Variant Compare At Price",
]


@dataclass
class Campaign:
    name: str
    column: str
    kind: str
    starts_at: str = ""
    ends_at: str = ""


def normalize(value: Any) -> str:
    text = "" if pd.isna(value) else str(value)
    text = text.strip().upper()
    text = (
        text.replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
        .replace("Ñ", "N")
    )
    return re.sub(r"\s+", " ", text)


def normalize_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", normalize(value))


def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_discount(value: Any) -> float | None:
    if value is None or pd.isna(value) or value == "":
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace("%", "").replace(",", ".")
        try:
            number = float(text)
        except ValueError:
            return None
    if number > 1:
        number = number / 100
    if number < 0 or number >= 1:
        return None
    return number


def to_money(value: Any) -> float | None:
    if value is None or pd.isna(value) or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "-", str(name)).strip()
    cleaned = re.sub(r"\s+", " ", cleaned) or "Descuento"
    cleaned = cleaned[:31].strip(" -")
    base = cleaned or "Descuento"
    suffix = 2
    while cleaned in used:
        tail = f" {suffix}"
        cleaned = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(cleaned)
    return cleaned


def find_header_row(path: Path, required: list[str], sheet_name: str | int = 0) -> int:
    preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=40, dtype=object)
    required_norm = {normalize(value) for value in required}
    for idx, row in preview.iterrows():
        row_values = {normalize(value) for value in row.tolist() if clean_text(value)}
        if required_norm.issubset(row_values):
            return int(idx)
    raise ValueError(f"No encontre encabezados requeridos: {', '.join(required)}")


def find_revenue_header_row(path: Path) -> int:
    preview = pd.read_excel(path, sheet_name=0, header=None, nrows=40, dtype=object)
    required_any = {
        "ID PRODUCTO",
        "SKU",
        "VARIANT SKU",
        "MODCOL",
        "COD MOD COL",
        "COD_MOD_COL",
        "MODELO COLOR",
        "MOD-COL",
    }
    for idx, row in preview.iterrows():
        row_values = {normalize(value) for value in row.tolist() if clean_text(value)}
        if row_values.intersection(required_any):
            return int(idx)
    raise ValueError("No encontre encabezados de Revenue: debe traer ID PRODUCTO o MODCOL.")


def first_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized = {normalize(col): col for col in df.columns}
    for candidate in candidates:
        found = normalized.get(normalize(candidate))
        if found is not None:
            return found
    return None


def read_matrixify(path: Path) -> pd.DataFrame:
    excel = pd.ExcelFile(path)
    sheet_name = "Products" if "Products" in excel.sheet_names else excel.sheet_names[0]
    header_row = find_header_row(path, MATRIXIFY_COLUMNS, sheet_name=sheet_name)
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row, dtype=object)
    df = df.dropna(how="all").copy()
    missing = [column for column in MATRIXIFY_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"El Matrixify no tiene columnas requeridas: {', '.join(missing)}")
    return df


def validate_matrixify_vendor(matrixify_df: pd.DataFrame, expected_vendor: str) -> tuple[bool, str]:
    if "Vendor" not in matrixify_df.columns:
        return True, "No pude validar Vendor porque el Matrixify no trae columna Vendor."
    vendors = {
        clean_text(value).lower()
        for value in matrixify_df["Vendor"].dropna().tolist()
        if clean_text(value)
    }
    expected = expected_vendor.lower()
    if vendors and expected not in vendors:
        return False, f"Vendor esperado: {expected_vendor}. Vendors encontrados: {', '.join(sorted(vendors)[:8])}."
    return True, ""


def extract_revenue_lookup_values(path: Path) -> tuple[list[str], list[str]]:
    header_row = find_revenue_header_row(path)
    df = pd.read_excel(path, sheet_name=0, header=header_row, dtype=object, usecols=lambda col: True)
    id_col = first_column(df, ["ID PRODUCTO", "SKU", "VARIANT SKU"])
    modcol_col = first_column(df, ["MODCOL", "COD MOD COL", "COD_MOD_COL", "MODELO COLOR", "MOD-COL"])
    ids = sorted({clean_text(value) for value in df[id_col].dropna().tolist() if clean_text(value)}) if id_col else []
    modcols = sorted({clean_text(value) for value in df[modcol_col].dropna().tolist() if clean_text(value)}) if modcol_col else []
    return ids, modcols


def extract_revenue_ids(path: Path) -> list[str]:
    ids, _modcols = extract_revenue_lookup_values(path)
    return ids


def read_revenue(
    path: Path,
    product_lookup: dict[str, Any] | None = None,
) -> tuple[pd.DataFrame, list[Campaign], str, str | None, str | None, list[str]]:
    header_row = find_revenue_header_row(path)
    raw_top = pd.read_excel(path, sheet_name=0, header=None, nrows=header_row, dtype=object)
    df = pd.read_excel(path, sheet_name=0, header=header_row, dtype=object).dropna(how="all").copy()

    id_col = first_column(df, ["ID PRODUCTO", "SKU", "VARIANT SKU"])
    modcol_col = first_column(df, ["MODCOL", "COD MOD COL", "COD_MOD_COL", "MODELO COLOR", "MOD-COL"])
    if not id_col and not modcol_col:
        raise ValueError("El Revenue debe tener ID PRODUCTO o MODCOL.")
    if not id_col:
        df["ID PRODUCTO"] = ""
        id_col = "ID PRODUCTO"
    brand_col = first_column(df, ["MARCA", "BRAND", "VENDOR"])
    product_lookup = product_lookup or {}
    by_id = product_lookup.get("by_id", product_lookup) if isinstance(product_lookup, dict) else {}
    by_modcol = product_lookup.get("by_modcol", {}) if isinstance(product_lookup, dict) else {}

    if modcol_col and by_modcol:
        expanded_rows: list[pd.Series] = []
        for _, row in df.iterrows():
            current_id = clean_text(row.get(id_col)) if id_col else ""
            modcol_value = clean_text(row.get(modcol_col))
            modcol_info = by_modcol.get(normalize_key(modcol_value), {})
            ids_from_modcol = modcol_info.get("ids", []) if isinstance(modcol_info, dict) else []
            if current_id or not ids_from_modcol:
                expanded_rows.append(row)
                continue
            for sku in ids_from_modcol:
                new_row = row.copy()
                new_row[id_col] = sku
                expanded_rows.append(new_row)
        if expanded_rows:
            df = pd.DataFrame(expanded_rows).reset_index(drop=True)

    if by_id or by_modcol:
        if not modcol_col:
            df["MODCOL"] = df[id_col].map(lambda value: by_id.get(normalize_key(value), {}).get("modcol", ""))
            modcol_col = "MODCOL"
        else:
            missing_modcol = df[modcol_col].map(clean_text) == ""
            df.loc[missing_modcol, modcol_col] = df.loc[missing_modcol, id_col].map(
                lambda value: by_id.get(normalize_key(value), {}).get("modcol", "")
            )

        if not brand_col:
            df["MARCA"] = df[id_col].map(lambda value: by_id.get(normalize_key(value), {}).get("brand", ""))
            if modcol_col:
                missing_brand = df["MARCA"].map(clean_text) == ""
                df.loc[missing_brand, "MARCA"] = df.loc[missing_brand, modcol_col].map(
                    lambda value: by_modcol.get(normalize_key(value), {}).get("brand", "")
                )
            brand_col = "MARCA"
        else:
            missing_brand = df[brand_col].map(clean_text) == ""
            df.loc[missing_brand, brand_col] = df.loc[missing_brand, id_col].map(
                lambda value: by_id.get(normalize_key(value), {}).get("brand", "")
            )
            if modcol_col:
                missing_brand = df[brand_col].map(clean_text) == ""
                df.loc[missing_brand, brand_col] = df.loc[missing_brand, modcol_col].map(
                    lambda value: by_modcol.get(normalize_key(value), {}).get("brand", "")
                )

    ignored = {
        "ID PRODUCTO",
        "SKU",
        "VARIANT SKU",
        "MODCOL",
        "COD MOD COL",
        "COD_MOD_COL",
        "MODELO",
        "COLOR",
        "MODELO COLOR",
        "MOD-COL",
        "MARCA",
        "BRAND",
        "VENDOR",
        "SITIO",
        "OBSERVACION",
        "OBSERVACIONES",
        "PRODUCTO",
        "DESCRIPCION",
        "DESCRIPTION",
    }
    discount_like = [
        col
        for col in df.columns
        if "DCTO" in normalize(col) or "DESCUENTO" in normalize(col)
    ]
    campaign_columns = discount_like or [col for col in df.columns if normalize(col) not in ignored]

    campaigns: list[Campaign] = []
    invalid_columns: list[str] = []
    for column in campaign_columns:
        if column == id_col or column == modcol_col or column == brand_col:
            continue
        col_pos = list(df.columns).index(column)
        starts_at = ""
        ends_at = ""
        if header_row >= 2 and col_pos < raw_top.shape[1]:
            starts_at = clean_text(raw_top.iat[header_row - 2, col_pos])
        if header_row >= 1 and col_pos < raw_top.shape[1]:
            ends_at = clean_text(raw_top.iat[header_row - 1, col_pos])
        name_parts = [part for part in [starts_at, ends_at, clean_text(column)] if part]
        name = " - ".join(name_parts) if name_parts else clean_text(column)
        kind = "resto_mes" if "RESTO" in normalize(name) or "ANT" in normalize(name) else "programado"
        if df[column].dropna().empty:
            invalid_columns.append(clean_text(column))
            continue
        campaigns.append(Campaign(name=name, column=column, kind=kind, starts_at=starts_at, ends_at=ends_at))

    if not campaigns:
        raise ValueError("No encontre columnas de descuento/campana en el Revenue.")
    return df, campaigns, id_col, modcol_col, brand_col, invalid_columns


def build_match_key(row: pd.Series, id_col: str, modcol_col: str | None) -> str:
    sku = clean_text(row.get(id_col))
    modcol = clean_text(row.get(modcol_col)) if modcol_col else ""
    if modcol:
        return f"{sku}|MODCOL:{modcol}"
    return sku


def split_match_key(match_key: str) -> tuple[str, str]:
    if "|MODCOL:" in match_key:
        sku, modcol = match_key.split("|MODCOL:", 1)
        return sku, modcol
    return match_key, ""


def preferred_missing_code(match_key: str) -> tuple[str, str, str]:
    sku, modcol = split_match_key(match_key)
    if modcol:
        return normalize_key(modcol), "MODCOL", modcol
    return normalize_key(sku), "SKU", sku


def build_matrixify_index(matrixify_df: pd.DataFrame) -> tuple[dict[str, str], dict[str, int], dict[str, str], str]:
    group_col = "ID" if "ID" in matrixify_df.columns else "Handle"
    working = matrixify_df.copy()
    working["_group_key"] = working[group_col].map(clean_text)
    missing_group = working["_group_key"] == ""
    working.loc[missing_group, "_group_key"] = working.loc[missing_group, "Variant SKU"].map(clean_text)

    sku_to_group: dict[str, str] = {}
    modcol_to_group: dict[str, str] = {}
    group_counts: dict[str, int] = working["_group_key"].value_counts().to_dict()

    for _, row in working.iterrows():
        group = clean_text(row["_group_key"])
        sku = clean_text(row.get("Variant SKU"))
        if sku:
            sku_to_group[sku] = group

        handle = clean_text(row.get("Handle"))
        if handle:
            parts = [part for part in handle.split("-") if part]
            for size in range(1, min(len(parts), 4) + 1):
                suffix = "-".join(parts[-size:])
                modcol_to_group.setdefault(normalize_key(suffix), group)

    matrixify_df["_group_key"] = working["_group_key"]
    return sku_to_group, group_counts, modcol_to_group, "_group_key"


def resolve_group(match_key: str, sku_to_group: dict[str, str], modcol_to_group: dict[str, str]) -> str | None:
    sku = match_key
    modcol = ""
    if "|MODCOL:" in match_key:
        sku, modcol = match_key.split("|MODCOL:", 1)
    if sku in sku_to_group:
        return sku_to_group[sku]
    if modcol:
        return modcol_to_group.get(normalize_key(modcol))
    return None


def build_discount_workbook(
    matrixify_path: Path,
    revenue_path: Path,
    output_path: Path,
    selected_brands: list[str] | None = None,
    product_lookup: dict[str, dict[str, str]] | None = None,
) -> dict[str, pd.DataFrame]:
    matrixify_df = read_matrixify(matrixify_path)
    revenue_df, campaigns, id_col, modcol_col, brand_col, invalid_columns = read_revenue(revenue_path, product_lookup)

    selected_norm = {normalize(brand) for brand in selected_brands or []}
    not_affected_rows: list[dict[str, Any]] = []
    if selected_norm and brand_col:
        brand_mask = revenue_df[brand_col].map(normalize).isin(selected_norm)
        for _, row in revenue_df.loc[~brand_mask].iterrows():
            code = build_match_key(row, id_col, modcol_col)
            if clean_text(row.get(id_col)):
                not_affected_rows.append(
                    {
                        "Codigo input": code,
                        "Marca input": clean_text(row.get(brand_col)),
                        "Motivo": "Marca fuera de la seleccion",
                    }
                )
        revenue_scope = revenue_df.loc[brand_mask].copy()
    else:
        revenue_scope = revenue_df.copy()

    sku_to_group, group_counts, modcol_to_group, group_col = build_matrixify_index(matrixify_df)
    scope_keys = [
        build_match_key(row, id_col, modcol_col)
        for _, row in revenue_scope.iterrows()
        if clean_text(row.get(id_col))
    ]
    scope_groups = {group for key in scope_keys if (group := resolve_group(key, sku_to_group, modcol_to_group))}
    missing_records: dict[tuple[str, str], dict[str, Any]] = {}
    for key in scope_keys:
        if resolve_group(key, sku_to_group, modcol_to_group):
            continue
        unique_key, code_type, code_value = preferred_missing_code(key)
        missing_records.setdefault(
            (code_type, unique_key),
            {
                "Carga": "Alcance input",
                "Tipo codigo": code_type,
                "Codigo no encontrado": code_value,
                "SKUs input asociados": 0,
                "Descuento": "",
                "Motivo": "No existe en Matrixify",
            },
        )
        missing_records[(code_type, unique_key)]["SKUs input asociados"] += 1

    summary_rows: list[dict[str, Any]] = []
    percent_rows: list[dict[str, Any]] = []
    missing_rows: list[dict[str, Any]] = list(missing_records.values())
    used_sheets: set[str] = set()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for campaign in campaigns:
            sheet_name = safe_sheet_name(campaign.name, used_sheets)
            discounts_by_group: dict[str, float] = {}
            invalid_discount_count = 0

            for _, row in revenue_scope.iterrows():
                match_key = build_match_key(row, id_col, modcol_col)
                discount = to_discount(row.get(campaign.column))
                if row.get(campaign.column) not in (None, "") and not pd.isna(row.get(campaign.column)) and discount is None:
                    invalid_discount_count += 1
                    continue
                if discount is None:
                    continue
                group = resolve_group(match_key, sku_to_group, modcol_to_group)
                if group:
                    discounts_by_group[group] = max(discounts_by_group.get(group, 0), discount)
                else:
                    unique_key, code_type, code_value = preferred_missing_code(match_key)
                    record_key = (sheet_name, code_type, unique_key)
                    existing = next(
                        (
                            row
                            for row in missing_rows
                            if row["Carga"] == sheet_name
                            and row["Tipo codigo"] == code_type
                            and normalize_key(row["Codigo no encontrado"]) == unique_key
                        ),
                        None,
                    )
                    if existing:
                        existing["SKUs input asociados"] += 1
                        existing["Descuento"] = max(to_discount(existing["Descuento"]) or 0, discount)
                    else:
                        missing_rows.append(
                            {
                                "Carga": sheet_name,
                                "Tipo codigo": code_type,
                                "Codigo no encontrado": code_value,
                                "SKUs input asociados": 1,
                                "Descuento": discount,
                                "Motivo": "No existe en Matrixify",
                            }
                        )

            out_df = matrixify_df[matrixify_df[group_col].isin(scope_groups)].copy()
            if out_df.empty:
                out_df = matrixify_df.head(0).copy()
            discount_series = out_df[group_col].map(discounts_by_group).fillna(0)
            base_price = out_df["Variant Compare At Price"].map(to_money)
            current_price = out_df["Variant Price"].map(to_money)
            original_price = base_price.where(base_price.notna(), current_price)

            new_price = original_price * (1 - discount_series)
            out_df["Variant Price"] = pd.Series(new_price.round(2).to_numpy(), index=out_df.index, dtype=object)
            out_df.loc[original_price.isna(), "Variant Price"] = out_df.loc[original_price.isna(), "Variant Price"]
            out_df["Variant Compare At Price"] = pd.Series([""] * len(out_df), index=out_df.index, dtype=object)
            discounted_mask = discount_series > 0
            out_df.loc[discounted_mask, "Variant Compare At Price"] = original_price.loc[discounted_mask].round(2).astype(object)

            final_df = out_df[MATRIXIFY_COLUMNS].copy()
            final_df.to_excel(writer, index=False, sheet_name=sheet_name)

            for discount, groups in pd.Series(discounts_by_group).groupby(pd.Series(discounts_by_group)).groups.items():
                group_list = list(groups)
                percent_rows.append(
                    {
                        "Carga": sheet_name,
                        "% Descuento": f"{discount:.0%}",
                        "Productos / modelo-color": len(group_list),
                        "Variantes Matrixify": sum(group_counts.get(group, 0) for group in group_list),
                    }
                )

            summary_rows.append(
                {
                    "Hoja": sheet_name,
                    "Tipo": "Resto del mes" if campaign.kind == "resto_mes" else "Programado",
                    "Columna Revenue": clean_text(campaign.column),
                    "Productos / modelo-color en archivo": len(scope_groups),
                    "Productos / modelo-color con descuento": len(discounts_by_group),
                    "Filas Matrixify generadas": len(final_df),
                    "Filas con descuento": int(discounted_mask.sum()),
                    "Filas sin descuento": int(len(final_df) - discounted_mask.sum()),
                    "Descuentos invalidos": invalid_discount_count,
                    "Inicio": campaign.starts_at,
                    "Fin": campaign.ends_at,
                }
            )

        summary_df = pd.DataFrame(summary_rows)
        percent_df = pd.DataFrame(percent_rows)
        missing_df = pd.DataFrame(missing_rows)
        not_affected_df = pd.DataFrame(not_affected_rows)
        invalid_columns_df = pd.DataFrame({"Columna sin datos": invalid_columns})

        summary_df.to_excel(writer, index=False, sheet_name="Resumen")
        percent_df.to_excel(writer, index=False, sheet_name="Descuentos por %")
        if not missing_df.empty:
            missing_df.to_excel(writer, index=False, sheet_name="No encontrados")
        if not not_affected_df.empty:
            not_affected_df.to_excel(writer, index=False, sheet_name="No afectados por marca")
        if not invalid_columns_df.empty:
            invalid_columns_df.to_excel(writer, index=False, sheet_name="Columnas sin datos")

    format_workbook(output_path)
    return {
        "summary": summary_df,
        "percent": percent_df,
        "missing": missing_df,
        "not_affected": not_affected_df,
    }


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(column_cells[0].value or "")) + 4, 14), 34)
        header_values = [cell.value for cell in sheet[1]]
        for price_column in ("Variant Price", "Variant Compare At Price"):
            if price_column in header_values:
                col_idx = header_values.index(price_column) + 1
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=col_idx).number_format = "0.00"
    workbook.save(path)
